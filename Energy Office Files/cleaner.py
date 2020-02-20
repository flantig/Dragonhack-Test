import pandas as pd
import openpyxl as pyxl
import csv

'''The stuff below is just direct excel-to-csv, it takes a whole hour to do'''
'''since I know it works, I'll probably never use this until the very end'''
'''this just means that if I decide to do any changes in regards to the headers in any excel file, testing it will be on a 1-by-one basis'''
'''the dictionary is monotonous'''

'''
#The intent for the openpyxl thing below is to just drag and drop the xlsm file and just run python to get the conversion to csv and cleaning done on its own.
wb = pyxl.load_workbook(r'S:\Energy\Meter Readings\Obvius SERC and HSC\Obvius Combined v1.1 Slimmed Down AH.xlsm', read_only=True)
wb.iso_dates = True
#read_only mode makes it so that it doesn't load up the excel sheet all at once and iso_dates tries to fix excel jank with the dates on their own
#spoilers...it doesn't really fix anything. So you had to go into the csvs below to go through the things that broke.

sheetnames = wb.get_sheet_names()
#get_sheet_names is depreciated but it seems to function the same as the new "sheetnames", which is to say, it returns a list of sheetnames
#I must have been reading some stackoverflow that used it before it was phased out. I'm going to modify this later

for nextsheetname in sheetnames:
    #Sifitng through all the workbook names
    sheet = wb[nextsheetname]
    #Note how we're using the string value or nextsheetname, this is because Python has a thing called pass-by-value
    #
    with open(r'S:\Energy\Database Progress\ObviusWorksheetsTest\%s.csv' %nextsheetname, 'w', newline= '') as f:
        c = csv.writer(f)
        #creating a writer object
        #this writes things naturally, but,
        for row in sheet.rows:
            values = (cell.value for cell in row)
            c.writerow(values)


    f.close()
#This process takes time...
#I think it takes like an hour? The plan is to automate absolutely everything and just let it run until everything is crammed into the database.
#As long as only the setup is slow and the database is fast, this is fine I guess
'''






filepath1 = r'S:\Energy\Database Progress\loggerlist.txt'
filepath2 = r'S:\Energy\Database Progress\meterfilename.txt'
logger_list = []
meterfilename = []
#file importing, so I don't clutter the code too much, would like to do the same with dtype but can't really figure it out atm

with open (filepath1, 'r') as fp1:
    logger_list = [line.strip().split(',') for line in fp1]
#Goes through every line in the file (for line in fp1) and says strip the newline and also divide them by comma into an array
#it makes 2d arrays, which is what you want


with open (filepath2, 'r') as fp2:
    meterfilename = fp2.readline().split(',')




def fixingDates(meterfile):
    '''So this is trying to resolve the horrifying mess that is excel.'''
    '''First, it keeps track of the position of Date and Time in any given spreadsheet just in case they happen to be in a different position for whatever reason'''
    testerino_data = open(r'S:\Energy\Database Progress\ObviusWorksheetsTest\\' + meterfile + '.csv', 'r')
    testerino = testerino_data.readlines()
    header = testerino[1].split(',')

    indexDate = 0
    indexTime = 0
    for i in header:
        if i == 'Date':
            break
        else:
            indexDate = indexDate + 1

    for j in header:
        if j == 'Time':
            break
        else:
            indexTime = indexTime + 1
    indexTime = indexTime + 1

    '''Second, it opens up whatever file we're trying to fix up '''
    with open(r'S:\Energy\Database Progress\ObviusWorksheetsTest\\' + meterfile + '.csv') as input, open(r'S:\Energy\Database Progress\FixingExcelsFormatting\\' + meterfile + '.csv', 'w', newline='') as cleaned:

        write = csv.writer(cleaned)
        initialfile = csv.reader(input)

        for c, row in enumerate(initialfile):
            if c==0:
                '''[:] allows you to select the entire row, which is what we're doing here. Other applications for this is selecting a range'''
                '''For example: row[0:2] starts from the first row index 0 and grabs the next two'''
                '''You didn't need to even put in a 0 to start from the first index [:2] accomplishes the same thing. The blank space means from the beginning of the list'''
                '''this is why [:] grabs everything. It starts from the beginning and goes all the way to the end.'''
                header=row[:]
                '''We're deleting the old header to fix it in later loops'''
                del header[indexDate]
                header[indexDate:indexDate] = ['Date', 'Leftovers']
                write.writerow(header)
            else:
                '''splitting junk and replacing symbols that shouldn't be'''
                date_reformater = [i.strip().replace('-', '/') for i in row[indexDate].split(' 00:00:00')]
                del row[indexDate]
                row[indexDate:indexDate] = date_reformater
                '''This very specific date is what excel seems to always default too when the time is 24 but it rally means to say 00 because the 23rd hour'''
                '''goes all the way up to 24 from which it then resets back to 00. This is convenient because if it weren't the same string'''
                '''this would be a pain to program but we just search for this specific string to fix'''
                if row[indexTime] == "1899-12-30 00:00:00":
                    row[indexTime] = "00:00:00"
                write.writerow(row)

    testerino_data.close()
    input.close()
    cleaned.close()




dtypes = [{"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.SWG1.INST (kW)": float, "SER.ELEC.SWG2.INST (kW)": float, "SER.ELEC.DHBA.INST (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SERC.CHW.TOTAL (kBtu)": int, "SERC.CHW.TOTAL Rate (kBtu/hr)": int, "SERC.STM.TOTAL (Lbs)": int, "SERC.STM.TOTAL Rate (Lb/hr)": int, "SER.HW.TOTAL (kBtu)": int, "SER.HW.TOTAL Rate (kBtu/hr)": int, "SER.PCHW.TOTAL (kBtu)": int, "SER.PCHW.TOTAL Rate (kBtu/hr)": int},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.DH1A.INST (kW)": float, "SER.ELEC.DH2A.INST (kW)": float, "SER.ELEC.DH3A.INST (kW)": float, "SER.ELEC.DH4A.INST (kW)": float, "SER.ELEC.DH5A.INST (kW)": float, "SER.ELEC.DH6A.INST (kW)": float, "SER.ELEC.DH7A.INST (kW)": float, "SER.ELEC.DH8A.INST (kW)": float}
          ]


pd.DataFrame(dtypes)

def cleanerObvius(dictionary, meterfile, dtypes):
        #df = pd.read_csv(r'S:\Energy\Database Progress\ObviusWorksheets\\' + meterfile[n] + '.csv', parse_dates=[['Date', 'Time']], skiprows=[1], usecols= dictionary[n])
        df = pd.read_csv(r'S:\Energy\Database Progress\FixingExcelsFormatting\\' + meterfile + '.csv', parse_dates=[['Date', 'Time']], skiprows=[1], usecols=dictionary)
        first_nan = df[df.Date_Time == 'nan nan'].first_valid_index()
        df = df[:first_nan]
        df.fillna(0)
        df = df.astype(dtypes)
        print(df)  # printing out the dataframe for testing purposes
        df.info()  # tells us what the columns are and their data type but I don't think it's necesarry at the moment beyond appeasing the compiler's warning
        df.to_csv(r'S:\Energy\Database Progress\CleanedWorksheets\\' + meterfile + '.csv', index=False)

for count, item in enumerate(dtypes):
    fixingDates(meterfilename[count])
    cleanerObvius(logger_list[count], meterfilename[count], item)

''' ================================================================================= '''
''' Pending Dictionary Below'''
''' ================================================================================= '''


'''

dtypes = [{"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.SWG1.INST (kW)": float, "SER.ELEC.SWG2.INST (kW)": float, "SER.ELEC.DHBA.INST (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SERC.CHW.TOTAL (kBtu)": int, "SERC.CHW.TOTAL Rate (kBtu/hr)": int, "SERC.STM.TOTAL (Lbs)": int, "SERC.STM.TOTAL Rate (Lb/hr)": int, "SER.HW.TOTAL (kBtu)": int, "SER.HW.TOTAL Rate (kBtu/hr)": int, "SER.PCHW.TOTAL (kBtu)": int, "SER.PCHW.TOTAL Rate (kBtu/hr)": int},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.DH1A.INST (kW)": float, "SER.ELEC.DH2A.INST (kW)": float, "SER.ELEC.DH3A.INST (kW)": float, "SER.ELEC.DH4A.INST (kW)": float, "SER.ELEC.DH5A.INST (kW)": float, "SER.ELEC.DH6A.INST (kW)": float, "SER.ELEC.DH7A.INST (kW)": float, "SER.ELEC.DH8A.INST (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-OD-1": float, "EL-OD-2201 (kWh)": float, "EL-OD-2201 Demand (kW)": float, "EL-OD-2207 (kWh)": float, "EL-OD-2207 Demand (kW)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-JH-1 (Lb/hr)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Heat Flow (KBtu/Min)": float, "Mass Flow (Lb/Min)": float, "STD Volume Flow (Gpm)": float, "Volume Flow (Gpm)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-KRE-1": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-MRB-1": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-SH-1 (Lb/hr)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-SFC-1 (Lb/hr)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Energy Consumption (kWh)": float, "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Total Real Power Present Demand (kW)": float, "Total Real Power Max Demand (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "BTU-BOY-1 (Btu)": float, "ST-BOY-1 (Lb/hr)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "BTU-BOY-1 (Btu)": float, "ST-BOY-1 (Lb/hr)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "ST-MOB-1": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Energy Rate (Btu/hr)": float, "Volume Rate (Gpm)": float, "Supply Temperature (F)": float, "Return Temperature (F)": float, "Energy Total Mode 1 (Btu)": float, "Energy Total Mode 2 (Btu)": float, "Volume Total Mode 1 (Gallons)": float, "Volume Total Mode 2 (Gallons)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "Real Power Phase A (kW)": float, "Real Power Phase B (kW)": float, "Real Power Phase C (kW)": float, "Energy Real In (kWh)": float, "Energy Real Out (kWh)": float, "Last Demand Real Power (kW)": float, "Present Demand Real Power (kW)": float, "Peak Demand Real Power (kW)": float},
          ]

'''




''' ================================================================================= '''
''' OLD CODE BELOW WITH DOCUMENTATION EXPLAINING ABOVE'''
''' ================================================================================= '''


'''tqdm is a progress bar so that I know which file ended up actually breaking'''


'''
SERC_Main_Electric_Device_3 = ['ID', 'TF Bldg', 'Date', 'Time', 'SER.ELEC.SWG1.INST (kW)', 'SER.ELEC.SWG2.INST (kW)', 'SER.ELEC.DHBA.INST (kW)']
dtypeTest1 = {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.SWG1.INST (kW)": float, "SER.ELEC.SWG2.INST (kW)": float, "SER.ELEC.DHBA.INST (kW)": float}
dtypeTest2 = {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.SWG1.INST (kW)": float, "SER.ELEC.SWG2.INST (kW)": float, "SER.ELEC.DHBA.INST (kW)": float}
#'M' converts the date_time from string AM/PM time to 24-Hour time which is exactly what we want without much fuss thankfully
df = pd.read_csv(r'S:\Energy\Database Progress\ObviusWorksheets\SERC-Main Electric Device 3.csv', parse_dates=[['Date', 'Time']], skiprows=[1], usecols = SERC_Main_Electric_Device_3)
#parse_dates takes the date and time cols and puts them together

first_nan = df[df.Date_Time =='nan nan'].first_valid_index()
# this gets the first occurence of nan nan or as represented in excel as an empty space and it specifically checks in the date_time row which will always have a value
last_index = df.last_valid_index()
#this gives us the last significant value


print("This is first occurence of nan nan " , first_nan)
print("This is last index of SERC main electric device 3 " , last_index)
#the two above just print out indexes for the first occurence of a NAN entry and the last entry ever

df = df[:first_nan]
#assigns the dataframe from the first element to the first blank element

df = df.astype(dtypeTest1)
#in case I ever want to do some number manipulation in pandas, I can because I correct the warning first issued when the program starts and assign a proper dtype to the columns


print(df) #printing out the dataframe for testing purposes
df.info() #tells us what the columns are and their data type but I don't think it's necesarry at the moment beyond appeasing the compiler's warning
df.to_csv(r'S:\Energy\Database Progress\CleanedWorksheets\SERC_Main_Electric_Device_3.csv', index=False)
'''

'''
logger_list = [['ID', 'TF Bldg', 'Date', 'Time', 'SER.ELEC.SWG1.INST (kW)', 'SER.ELEC.SWG2.INST (kW)', 'SER.ELEC.DHBA.INST (kW)'],
               ['ID', 'TF Bldg', 'Date', 'Time', 'SERC.CHW.TOTAL (kBtu)', 'SERC.CHW.TOTAL Rate (kBtu/hr)', 'SERC.STM.TOTAL (Lbs)', 'SERC.STM.TOTAL Rate (Lb/hr)', 'SER.HW.TOTAL (kBtu)', 'SER.HW.TOTAL Rate (kBtu/hr)', 'SER.PCHW.TOTAL (kBtu)', 'SER.PCHW.TOTAL Rate (kBtu/hr)'],
               ['ID', 'TF Bldg', 'Date', 'Time', 'SER.ELEC.DH1A.INST (kW)', 'SER.ELEC.DH2A.INST (kW)', 'SER.ELEC.DH3A.INST (kW)', 'SER.ELEC.DH4A.INST (kW)', 'SER.ELEC.DH5A.INST (kW)', 'SER.ELEC.DH6A.INST (kW)', 'SER.ELEC.DH7A.INST (kW)', 'SER.ELEC.DH8A.INST (kW)']]
'''

'''
meterfilename = ['SERC-Main Electric Device 3',
                 'SERC-Stm and CHW Bldg Device 4',
                 'SERC-Floor Electric Device 1']
'''

'''
dtypes = [{"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.SWG1.INST (kW)": float, "SER.ELEC.SWG2.INST (kW)": float, "SER.ELEC.DHBA.INST (kW)": float},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SERC.CHW.TOTAL (kBtu)": int, "SERC.CHW.TOTAL Rate (kBtu/hr)": int, "SERC.STM.TOTAL (Lbs)": int, "SERC.STM.TOTAL Rate (Lb/hr)": int, "SER.HW.TOTAL (kBtu)": int, "SER.HW.TOTAL Rate (kBtu/hr)": int, "SER.PCHW.TOTAL (kBtu)": int, "SER.PCHW.TOTAL Rate (kBtu/hr)": int},
          {"ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.DH1A.INST (kW)": float, "SER.ELEC.DH2A.INST (kW)": float, "SER.ELEC.DH3A.INST (kW)": float, "SER.ELEC.DH4A.INST (kW)": float, "SER.ELEC.DH5A.INST (kW)": float, "SER.ELEC.DH6A.INST (kW)": float, "SER.ELEC.DH7A.INST (kW)": float, "SER.ELEC.DH8A.INST (kW)": float}]
'''

'''
dtypes = [["ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.SWG1.INST (kW)": float, "SER.ELEC.SWG2.INST (kW)": float, "SER.ELEC.DHBA.INST (kW)": float],
          ["ID": str, "TF Bldg": str, "Date_Time": 'M', "SERC.CHW.TOTAL (kBtu)": int, "SERC.CHW.TOTAL Rate (kBtu/hr)": int, "SERC.STM.TOTAL (Lbs)": int, "SERC.STM.TOTAL Rate (Lb/hr)": int, "SER.HW.TOTAL (kBtu)": int, "SER.HW.TOTAL Rate (kBtu/hr)": int, "SER.PCHW.TOTAL (kBtu)": int, "SER.PCHW.TOTAL Rate (kBtu/hr)": int],
          ["ID": str, "TF Bldg": str, "Date_Time": 'M', "SER.ELEC.DH1A.INST (kW)": float, "SER.ELEC.DH2A.INST (kW)": float, "SER.ELEC.DH3A.INST (kW)": float, "SER.ELEC.DH4A.INST (kW)": float, "SER.ELEC.DH5A.INST (kW)": float, "SER.ELEC.DH6A.INST (kW)": float, "SER.ELEC.DH7A.INST (kW)": float, "SER.ELEC.DH8A.INST (kW)": float]]
'''

